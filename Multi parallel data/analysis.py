#!/usr/bin/env python3
"""PPO Flappy Bird Training Analysis Dashboard.

Loads 60 Excel files across 5 run folders, computes key metrics,
and serves an interactive Dash dashboard on http://localhost:8050.
"""

import os
import re
import ast
import warnings

import numpy as np
import pandas as pd
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import dash
from dash import dcc, html, dash_table, callback_context
from dash.dependencies import Input, Output, State
import dash_bootstrap_components as dbc

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DATA_DIR = os.path.dirname(os.path.abspath(__file__))

# Raw architecture strings -> short human labels
# These match the strings stored in Config sheets
ARCH_LABELS = {
    "[[8, 1024], [1024, 512], [512, 2048]]":                                       "M0: 1024-512-2048",
    "[[8, 2048], [2048, 512], [512, 2048]]":                                        "M1: 2048-512-2048",
    "[[8, 1024], [1024, 512], [512, 256], [256, 256], [256, 512], [512, 1024]]":    "M2: 1024-512-256-256-512-1024",
    "[[8, 1024], [1024, 128], [128, 16], [16, 128], [128, 256]]":                   "M3: 1024-128-16-128-256",
    "[[8, 512], [512, 64], [64, 256], [256, 256], [256, 256]]":                     "M4: 512-64-256-256-256",
    "[[8, 16], [16, 8], [8, 16], [16, 8], [8, 16]]":                               "M5: 16-8-16-8-16",
    "[[8, 1024], [1024, 8], [8, 256], [256, 256]]":                                 "M6: 1024-8-256-256",
    "[[8, 1024], [1024, 128], [128, 16], [16, 256], [256, 256]]":                   "M7: 1024-128-16-256-256",
    "[[8, 64], [64, 128], [128, 128]]":                                             "M8: 64-128-128",
    "[[8, 16], [16, 1024], [1024, 256]]":                                           "M9: 16-1024-256",
}

MODEL_COLORS = {
    "M0: 1024-512-2048":            "#00d4ff",
    "M1: 2048-512-2048":            "#ff6b6b",
    "M2: 1024-512-256-256-512-1024":"#51cf66",
    "M3: 1024-128-16-128-256":      "#ffd43b",
    "M4: 512-64-256-256-256":       "#cc5de8",
    "M5: 16-8-16-8-16":             "#ff922b",
    "M6: 1024-8-256-256":           "#20c997",
    "M7: 1024-128-16-256-256":      "#339af0",
    "M8: 64-128-128":               "#f06595",
    "M9: 16-1024-256":              "#adb5bd",
}

DARK_BG = "#1a1a2e"
CARD_BG = "#16213e"
ACCENT = "#00d4ff"

RUN_FOLDER_MAP = {
    "Run 1 Lr":       "Run 1: Learning Rate",
    "Run 2 Lr":       "Run 2: Learning Rate",
    "Run 3 K Epochs": "Run 3: K Epochs",
    "Run 4 Epsilon":  "Run 4: Epsilon",
    "Run 5 gamma":    "Run 5: Gamma",
}

SWEPT_PARAM = {
    "Run 1 Lr":       "lr",
    "Run 2 Lr":       "lr",
    "Run 3 K Epochs": "K_epochs",
    "Run 4 Epsilon":  "epsilon",
    "Run 5 gamma":    "gamma",
}


def compute_param_count(arch_str: str) -> int:
    """Compute trainable parameter count for a given architecture string."""
    layers = ast.literal_eval(arch_str)
    total = 0
    for in_dim, out_dim in layers:
        total += in_dim * out_dim + out_dim  # weights + biases
    return total


ARCH_PARAMS = {}
for _raw, _label in ARCH_LABELS.items():
    ARCH_PARAMS[_label] = compute_param_count(_raw)


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------
def load_all_data() -> pd.DataFrame:
    """Load all Excel files and return a single DataFrame with all epoch data."""
    all_rows = []

    for folder_name in sorted(os.listdir(DATA_DIR)):
        folder_path = os.path.join(DATA_DIR, folder_name)
        if not os.path.isdir(folder_path) or folder_name not in RUN_FOLDER_MAP:
            continue

        run_group = folder_name

        for fname in sorted(os.listdir(folder_path)):
            if not fname.endswith(".xlsx"):
                continue
            fpath = os.path.join(folder_path, fname)

            # Parse config and run indices from filename
            m = re.match(r"config(\d+)_run(\d+)", fname)
            if not m:
                print(f"  Skipping unrecognized file: {fname}")
                continue
            config_idx = int(m.group(1))
            run_idx = int(m.group(2))

            try:
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            except Exception as e:
                print(f"  WARNING: Could not open {fpath}: {e}")
                continue

            # --- Read Config sheet ---
            try:
                ws_cfg = wb["Config"]
                cfg_rows = list(ws_cfg.iter_rows(values_only=True))
            except Exception as e:
                print(f"  WARNING: No Config sheet in {fpath}: {e}")
                wb.close()
                continue

            hyperparams = {}
            arch_map = {}  # model_index -> arch_label
            for row in cfg_rows:
                if row is None or len(row) < 2:
                    continue
                key, val = row[0], row[1]
                if key is None:
                    continue
                key_str = str(key).strip()

                if key_str in ("lr", "epochs", "K_epochs", "epsilon", "gamma", "c0", "c1", "c2"):
                    hyperparams[key_str] = float(val) if val is not None else None
                elif "architecture" in key_str.lower():
                    idx_match = re.search(r"(\d+)", key_str)
                    if idx_match and val is not None:
                        midx = int(idx_match.group(1))
                        raw_arch = str(val).strip()
                        label = ARCH_LABELS.get(raw_arch)
                        if label:
                            arch_map[midx] = label
                        else:
                            arch_map[midx] = f"Unknown({raw_arch[:30]})"

            # --- Read Model sheets ---
            model_sheets = [s for s in wb.sheetnames if s.startswith("Model")]
            for sheet_name in model_sheets:
                idx_match = re.search(r"(\d+)", sheet_name)
                if not idx_match:
                    continue
                model_idx = int(idx_match.group(1))
                arch_label = arch_map.get(model_idx, f"Model {model_idx}")

                try:
                    ws_model = wb[sheet_name]
                    first = True
                    for row in ws_model.iter_rows(values_only=True):
                        if first:
                            first = False
                            continue  # skip header
                        if row is None or len(row) < 5:
                            continue
                        epoch, reward, l_clip, l_vf, l_entropy = row[:5]
                        if epoch is None:
                            continue
                        all_rows.append({
                            "run_group":  run_group,
                            "config_idx": config_idx,
                            "run_idx":    run_idx,
                            "arch_label": arch_label,
                            "epoch":      int(epoch),
                            "reward":     float(reward) if reward is not None else np.nan,
                            "L_clip":     float(l_clip) if l_clip is not None else np.nan,
                            "L_vf":       float(l_vf) if l_vf is not None else np.nan,
                            "L_entropy":  float(l_entropy) if l_entropy is not None else np.nan,
                            "lr":         hyperparams.get("lr"),
                            "K_epochs":   hyperparams.get("K_epochs"),
                            "epsilon":    hyperparams.get("epsilon"),
                            "gamma":      hyperparams.get("gamma"),
                            "c1":         hyperparams.get("c1"),
                            "c2":         hyperparams.get("c2"),
                        })
                except Exception as e:
                    print(f"  WARNING: Error reading {sheet_name} in {fpath}: {e}")

            wb.close()

    df = pd.DataFrame(all_rows)
    print(f"Loaded {len(df):,} rows from {df[['run_group','config_idx','run_idx']].drop_duplicates().shape[0]} files")
    return df


# ---------------------------------------------------------------------------
# Metrics Computation
# ---------------------------------------------------------------------------
def compute_summary(df: pd.DataFrame) -> pd.DataFrame:
    """One row per (run_group, config_idx, run_idx, arch_label)."""
    group_cols = ["run_group", "config_idx", "run_idx", "arch_label"]

    def agg_fn(g):
        rewards = g["reward"].values
        return pd.Series({
            "best_reward":       np.nanmax(rewards),
            "avg_last_100":      np.nanmean(rewards[-100:]),
            "first_epoch_to_zero": int(g.loc[g["reward"] >= 0, "epoch"].min())
                                   if (g["reward"] >= 0).any() else np.nan,
            "mean_reward":       np.nanmean(rewards),
            "lr":                g["lr"].iloc[0],
            "K_epochs":          g["K_epochs"].iloc[0],
            "epsilon":           g["epsilon"].iloc[0],
            "gamma":             g["gamma"].iloc[0],
            "c1":                g["c1"].iloc[0],
            "c2":                g["c2"].iloc[0],
        })

    summary = df.groupby(group_cols, sort=False).apply(agg_fn, include_groups=False).reset_index()
    return summary


def compute_kpis(summary: pd.DataFrame) -> dict:
    """Top-level KPI values for the Overview tab."""
    best_row = summary.loc[summary["best_reward"].idxmax()]
    best_avg_row = summary.loc[summary["avg_last_100"].idxmax()]

    arch_means = summary.groupby("arch_label")["avg_last_100"].mean()
    best_arch = arch_means.idxmax()

    valid_ftz = summary.dropna(subset=["first_epoch_to_zero"])
    if len(valid_ftz) > 0:
        fastest_row = valid_ftz.loc[valid_ftz["first_epoch_to_zero"].idxmin()]
        arch_ftz = valid_ftz.groupby("arch_label")["first_epoch_to_zero"].mean()
        fastest_arch = arch_ftz.idxmin()
        fastest_arch_val = arch_ftz.min()
    else:
        fastest_row = None
        fastest_arch = "N/A"
        fastest_arch_val = np.nan

    success_rate = (summary["first_epoch_to_zero"].notna().sum() / len(summary) * 100)

    return {
        "best_reward":       best_row["best_reward"],
        "best_reward_id":    f"{best_row['arch_label']} ({best_row['run_group']}, c{int(best_row['config_idx'])}r{int(best_row['run_idx'])})",
        "best_avg":          best_avg_row["avg_last_100"],
        "best_avg_id":       f"{best_avg_row['arch_label']} ({best_avg_row['run_group']}, c{int(best_avg_row['config_idx'])}r{int(best_avg_row['run_idx'])})",
        "best_arch":         best_arch,
        "best_arch_avg":     arch_means[best_arch],
        "fastest_single":    fastest_row["first_epoch_to_zero"] if fastest_row is not None else np.nan,
        "fastest_single_id": f"{fastest_row['arch_label']} ({fastest_row['run_group']})" if fastest_row is not None else "N/A",
        "fastest_arch":      fastest_arch,
        "fastest_arch_avg":  fastest_arch_val,
        "success_rate":      success_rate,
        "total_experiments": len(summary),
    }


def compute_arch_comparison(summary: pd.DataFrame) -> pd.DataFrame:
    """Per-architecture aggregated stats."""
    def agg_arch(g):
        return pd.Series({
            "mean_best_reward":  g["best_reward"].mean(),
            "mean_avg_last_100": g["avg_last_100"].mean(),
            "mean_first_to_zero": g["first_epoch_to_zero"].mean(),
            "param_count":       ARCH_PARAMS.get(g.name, 0),
            "success_rate":      g["first_epoch_to_zero"].notna().mean() * 100,
            "n_experiments":     len(g),
        })

    arch_comp = summary.groupby("arch_label").apply(agg_arch, include_groups=False).reset_index()
    arch_comp = arch_comp.sort_values("mean_avg_last_100", ascending=False)
    return arch_comp


def compute_leaderboard(summary: pd.DataFrame) -> pd.DataFrame:
    """All experiments ranked by avg_last_100."""
    lb = summary.copy()
    lb["param_count"] = lb["arch_label"].map(ARCH_PARAMS)
    lb["rank"] = lb["avg_last_100"].rank(ascending=False, method="min").astype(int)
    lb = lb.sort_values("rank")
    return lb


def compute_all(df):
    summary = compute_summary(df)
    kpis = compute_kpis(summary)
    arch_comp = compute_arch_comparison(summary)
    leaderboard = compute_leaderboard(summary)
    return summary, kpis, arch_comp, leaderboard


# ---------------------------------------------------------------------------
# Chart Helpers
# ---------------------------------------------------------------------------
def dark_layout(fig, title="", height=500):
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor=DARK_BG,
        plot_bgcolor=DARK_BG,
        title=dict(text=title, font=dict(size=16, color="white")),
        font=dict(color="#e0e0e0"),
        height=height,
        margin=dict(l=60, r=30, t=50, b=50),
        legend=dict(bgcolor="rgba(0,0,0,0)"),
    )
    return fig


# ---------------------------------------------------------------------------
# Chart Functions
# ---------------------------------------------------------------------------
def fig_arch_performance(arch_comp):
    ac = arch_comp.sort_values("mean_avg_last_100", ascending=True)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=ac["arch_label"], x=ac["mean_avg_last_100"],
        orientation="h",
        marker_color=[MODEL_COLORS.get(a, "#888") for a in ac["arch_label"]],
        text=[f"{v:.1f}" for v in ac["mean_avg_last_100"]],
        textposition="outside",
    ))
    dark_layout(fig, "Architecture Performance (Mean Avg Last 100 Epochs)", height=450)
    fig.update_layout(xaxis_title="Mean Reward (Last 100 Epochs)", yaxis_title="")
    return fig


def fig_param_vs_perf(arch_comp):
    fig = go.Figure()
    for _, row in arch_comp.iterrows():
        label = row["arch_label"]
        fig.add_trace(go.Scatter(
            x=[row["param_count"]],
            y=[row["mean_avg_last_100"]],
            mode="markers+text",
            marker=dict(size=14, color=MODEL_COLORS.get(label, "#888")),
            text=[label.split(":")[0]],
            textposition="top center",
            name=label,
            showlegend=False,
        ))
    dark_layout(fig, "Parameter Count vs Performance", height=400)
    fig.update_layout(xaxis_title="Trainable Parameters", yaxis_title="Mean Reward (Last 100)")
    fig.update_xaxes(type="log")
    return fig


def fig_success_rate(arch_comp):
    ac = arch_comp.sort_values("success_rate", ascending=True)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=ac["arch_label"], x=ac["success_rate"],
        orientation="h",
        marker_color=[MODEL_COLORS.get(a, "#888") for a in ac["arch_label"]],
        text=[f"{v:.0f}%" for v in ac["success_rate"]],
        textposition="outside",
    ))
    dark_layout(fig, "Success Rate by Architecture (% Reaching Reward >= 0)", height=450)
    fig.update_layout(xaxis_title="Success Rate (%)", yaxis_title="", xaxis_range=[0, 105])
    return fig


def fig_hyperparam_sweep(summary, run_group):
    """Line chart with error bands for the swept parameter in a run group."""
    sub = summary[summary["run_group"] == run_group].copy()
    if sub.empty:
        return dark_layout(go.Figure(), "No data")

    param = SWEPT_PARAM.get(run_group, "lr")
    sub["param_val"] = sub[param]

    agg = sub.groupby(["arch_label", "param_val"]).agg(
        mean_reward=("avg_last_100", "mean"),
        std_reward=("avg_last_100", "std"),
        count=("avg_last_100", "count"),
    ).reset_index()
    agg["std_reward"] = agg["std_reward"].fillna(0)

    fig = go.Figure()
    for arch in sorted(agg["arch_label"].unique()):
        a = agg[agg["arch_label"] == arch].sort_values("param_val")
        color = MODEL_COLORS.get(arch, "#888")
        fig.add_trace(go.Scatter(
            x=a["param_val"], y=a["mean_reward"],
            mode="lines+markers", name=arch,
            line=dict(color=color, width=2),
            marker=dict(size=6),
        ))

    dark_layout(fig, f"{RUN_FOLDER_MAP.get(run_group, run_group)}: Sweep Results", height=500)
    fig.update_layout(
        xaxis_title=param,
        yaxis_title="Mean Reward (Last 100 Epochs)",
    )
    if param == "lr":
        fig.update_xaxes(type="log")
    return fig


def fig_heatmap(summary, run_group):
    """Heatmap: architecture vs swept param value -> mean reward."""
    sub = summary[summary["run_group"] == run_group].copy()
    if sub.empty:
        return dark_layout(go.Figure(), "No data")

    param = SWEPT_PARAM.get(run_group, "lr")
    sub["param_val"] = sub[param]

    pivot = sub.pivot_table(
        values="avg_last_100", index="arch_label", columns="param_val", aggfunc="mean"
    )
    pivot = pivot.reindex(sorted(pivot.index))

    fig = go.Figure(data=go.Heatmap(
        z=pivot.values,
        x=[str(c) for c in pivot.columns],
        y=pivot.index,
        colorscale="Viridis",
        colorbar=dict(title="Avg Reward"),
        text=np.round(pivot.values, 1),
        texttemplate="%{text}",
        textfont=dict(size=10),
    ))
    dark_layout(fig, f"Reward Heatmap: Architecture vs {param}", height=500)
    fig.update_layout(xaxis_title=param, yaxis_title="")
    return fig


def fig_run_box(summary, run_group):
    """Box plot per swept param value within a run group."""
    sub = summary[summary["run_group"] == run_group].copy()
    if sub.empty:
        return dark_layout(go.Figure(), "No data")

    param = SWEPT_PARAM.get(run_group, "lr")
    sub["param_val"] = sub[param].astype(str)

    fig = px.box(
        sub, x="param_val", y="avg_last_100", color="arch_label",
        color_discrete_map=MODEL_COLORS,
    )
    dark_layout(fig, f"Reward Distribution by {param}", height=450)
    fig.update_layout(xaxis_title=param, yaxis_title="Avg Reward (Last 100)")
    return fig


def fig_model_box(summary):
    """Box plot of avg_last_100 by architecture across all runs."""
    fig = px.box(
        summary, x="arch_label", y="avg_last_100",
        color="arch_label", color_discrete_map=MODEL_COLORS,
        category_orders={"arch_label": sorted(summary["arch_label"].unique())},
    )
    dark_layout(fig, "Reward Distribution by Architecture (All Runs)", height=500)
    fig.update_layout(xaxis_title="", yaxis_title="Avg Reward (Last 100)", showlegend=False)
    fig.update_xaxes(tickangle=25)
    return fig


def fig_convergence_speed(arch_comp):
    """Bar chart of mean first_epoch_to_zero by architecture."""
    ac = arch_comp.dropna(subset=["mean_first_to_zero"]).sort_values("mean_first_to_zero")
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=ac["arch_label"], y=ac["mean_first_to_zero"],
        marker_color=[MODEL_COLORS.get(a, "#888") for a in ac["arch_label"]],
        text=[f"{v:.0f}" for v in ac["mean_first_to_zero"]],
        textposition="outside",
    ))
    dark_layout(fig, "Mean Epochs to Reach Reward >= 0", height=450)
    fig.update_layout(xaxis_title="", yaxis_title="Epochs to Zero Reward")
    fig.update_xaxes(tickangle=25)
    return fig


def fig_learning_curves(df, arch_labels, run_group=None, config_idx=None,
                        smoothing=50, show_individual=False):
    """Epoch-by-epoch reward plot with optional smoothing."""
    sub = df.copy()
    if run_group:
        sub = sub[sub["run_group"] == run_group]
    if config_idx is not None:
        sub = sub[sub["config_idx"] == config_idx]
    sub = sub[sub["arch_label"].isin(arch_labels)]

    fig = go.Figure()
    for arch in sorted(arch_labels):
        a = sub[sub["arch_label"] == arch]
        color = MODEL_COLORS.get(arch, "#888")

        if show_individual:
            for (rg, ci, ri), trial in a.groupby(["run_group", "config_idx", "run_idx"]):
                trial = trial.sort_values("epoch")
                smoothed = trial["reward"].rolling(window=smoothing, min_periods=1).mean()
                fig.add_trace(go.Scatter(
                    x=trial["epoch"], y=smoothed,
                    mode="lines", name=f"{arch} ({rg} c{ci}r{ri})",
                    line=dict(color=color, width=1),
                    opacity=0.4, showlegend=False,
                ))

        # Mean across all matching trials
        mean_curve = a.groupby("epoch")["reward"].mean().reset_index().sort_values("epoch")
        if smoothing > 1:
            mean_curve["reward"] = mean_curve["reward"].rolling(window=smoothing, min_periods=1).mean()
        fig.add_trace(go.Scatter(
            x=mean_curve["epoch"], y=mean_curve["reward"],
            mode="lines", name=arch,
            line=dict(color=color, width=2.5),
        ))

    dark_layout(fig, "Learning Curves", height=550)
    fig.update_layout(xaxis_title="Epoch", yaxis_title="Reward")
    return fig


def fig_loss_components(df, arch_label, run_group=None, config_idx=None, smoothing=50):
    """3-panel subplot of L_clip, L_vf, L_entropy over epochs."""
    sub = df[df["arch_label"] == arch_label].copy()
    if run_group:
        sub = sub[sub["run_group"] == run_group]
    if config_idx is not None:
        sub = sub[sub["config_idx"] == config_idx]

    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                        subplot_titles=["L_clip (Policy Loss)", "L_vf (Value Loss)", "L_entropy"],
                        vertical_spacing=0.08)

    mean_curve = sub.groupby("epoch")[["L_clip", "L_vf", "L_entropy"]].mean().reset_index().sort_values("epoch")
    for col in ["L_clip", "L_vf", "L_entropy"]:
        mean_curve[col] = mean_curve[col].rolling(window=smoothing, min_periods=1).mean()

    color = MODEL_COLORS.get(arch_label, ACCENT)
    for i, col in enumerate(["L_clip", "L_vf", "L_entropy"], 1):
        fig.add_trace(go.Scatter(
            x=mean_curve["epoch"], y=mean_curve[col],
            mode="lines", line=dict(color=color, width=2),
            name=col, showlegend=(i == 1),
        ), row=i, col=1)

    dark_layout(fig, f"Loss Components: {arch_label}", height=700)
    fig.update_xaxes(title_text="Epoch", row=3, col=1)
    return fig


def fig_loss_vs_reward(df, arch_label, run_group=None, config_idx=None):
    """Scatter: L_clip vs reward colored by epoch."""
    sub = df[df["arch_label"] == arch_label].copy()
    if run_group:
        sub = sub[sub["run_group"] == run_group]
    if config_idx is not None:
        sub = sub[sub["config_idx"] == config_idx]

    # Sample if too many points
    if len(sub) > 5000:
        sub = sub.sample(5000, random_state=42)

    fig = px.scatter(
        sub, x="L_clip", y="reward", color="epoch",
        color_continuous_scale="Viridis", opacity=0.5,
    )
    dark_layout(fig, f"Policy Loss vs Reward: {arch_label}", height=450)
    fig.update_layout(xaxis_title="L_clip", yaxis_title="Reward")
    return fig


# ---------------------------------------------------------------------------
# Dashboard Layout
# ---------------------------------------------------------------------------
def make_kpi_card(title, value, subtitle=""):
    return dbc.Card(
        dbc.CardBody([
            html.H6(title, className="text-muted mb-1", style={"fontSize": "0.85rem"}),
            html.H3(
                f"{value:.1f}" if isinstance(value, (int, float)) and not np.isnan(value) else str(value),
                style={"color": ACCENT, "fontWeight": "bold", "marginBottom": "0.2rem"},
            ),
            html.Small(subtitle, className="text-muted") if subtitle else html.Div(),
        ]),
        style={"backgroundColor": CARD_BG, "border": "1px solid #2a2a4a"},
        className="h-100",
    )


def build_app(df, summary, kpis, arch_comp, leaderboard):
    app = dash.Dash(
        __name__,
        external_stylesheets=[dbc.themes.DARKLY],
        suppress_callback_exceptions=True,
    )

    all_archs = sorted(df["arch_label"].unique())
    all_run_groups = sorted(df["run_group"].unique())

    # -----------------------------------------------------------------------
    # Overview Tab
    # -----------------------------------------------------------------------
    overview_tab = dbc.Container([
        dbc.Row([
            dbc.Col(make_kpi_card("Best All-Time Reward", kpis["best_reward"], kpis["best_reward_id"]), md=2),
            dbc.Col(make_kpi_card("Best Avg (Last 100)", kpis["best_avg"], kpis["best_avg_id"]), md=2),
            dbc.Col(make_kpi_card("Best Architecture", 0, kpis["best_arch"]), md=2),
            dbc.Col(make_kpi_card("Fastest to 0 (Avg)", kpis["fastest_arch_avg"],
                                  kpis["fastest_arch"]), md=2),
            dbc.Col(make_kpi_card("Success Rate", kpis["success_rate"], "% reaching reward >= 0"), md=2),
            dbc.Col(make_kpi_card("Total Experiments", kpis["total_experiments"], "across 5 runs"), md=2),
        ], className="mb-4 g-3"),
        dbc.Row([
            dbc.Col(dcc.Graph(figure=fig_arch_performance(arch_comp)), md=6),
            dbc.Col(dcc.Graph(figure=fig_param_vs_perf(arch_comp)), md=6),
        ], className="mb-3"),
        dbc.Row([
            dbc.Col(dcc.Graph(figure=fig_success_rate(arch_comp)), md=12),
        ]),
    ], fluid=True, className="py-3")

    # -----------------------------------------------------------------------
    # Per-Run Tab
    # -----------------------------------------------------------------------
    perrun_tab = dbc.Container([
        dbc.Row([
            dbc.Col([
                html.Label("Select Run Group", className="text-muted"),
                dcc.Dropdown(
                    id="perrun-run-group",
                    options=[{"label": RUN_FOLDER_MAP.get(rg, rg), "value": rg} for rg in all_run_groups],
                    value=all_run_groups[0] if all_run_groups else None,
                    className="mb-3",
                    style={"color": "#000"},
                ),
            ], md=4),
        ]),
        dbc.Row([
            dbc.Col(dcc.Graph(id="perrun-sweep"), md=12),
        ], className="mb-3"),
        dbc.Row([
            dbc.Col(dcc.Graph(id="perrun-heatmap"), md=6),
            dbc.Col(dcc.Graph(id="perrun-box"), md=6),
        ]),
    ], fluid=True, className="py-3")

    # -----------------------------------------------------------------------
    # Model Comparison Tab
    # -----------------------------------------------------------------------
    arch_table_data = []
    for _, r in arch_comp.iterrows():
        arch_table_data.append({
            "Architecture": r["arch_label"],
            "Parameters": f"{r['param_count']:,.0f}",
            "Mean Best Reward": f"{r['mean_best_reward']:.1f}",
            "Mean Avg (Last 100)": f"{r['mean_avg_last_100']:.1f}",
            "Avg Epochs to 0": f"{r['mean_first_to_zero']:.0f}" if not np.isnan(r["mean_first_to_zero"]) else "N/A",
            "Success Rate": f"{r['success_rate']:.0f}%",
            "N Experiments": int(r["n_experiments"]),
        })

    model_tab = dbc.Container([
        dbc.Row([
            dbc.Col(dcc.Graph(figure=fig_model_box(summary)), md=7),
            dbc.Col(dcc.Graph(figure=fig_convergence_speed(arch_comp)), md=5),
        ], className="mb-3"),
        dbc.Row([
            dbc.Col([
                html.H5("Architecture Details", className="text-muted mb-2"),
                dash_table.DataTable(
                    data=arch_table_data,
                    columns=[{"name": c, "id": c} for c in arch_table_data[0].keys()],
                    style_table={"overflowX": "auto"},
                    style_header={
                        "backgroundColor": CARD_BG, "color": "white",
                        "fontWeight": "bold", "border": "1px solid #2a2a4a",
                    },
                    style_cell={
                        "backgroundColor": DARK_BG, "color": "#e0e0e0",
                        "border": "1px solid #2a2a4a", "textAlign": "center",
                        "padding": "8px",
                    },
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#1e1e3a"},
                    ],
                    sort_action="native",
                ),
            ], md=12),
        ]),
    ], fluid=True, className="py-3")

    # -----------------------------------------------------------------------
    # Learning Curves Tab
    # -----------------------------------------------------------------------
    lc_tab = dbc.Container([
        dbc.Row([
            dbc.Col([
                html.Label("Architectures", className="text-muted"),
                dcc.Dropdown(
                    id="lc-archs",
                    options=[{"label": a, "value": a} for a in all_archs],
                    value=all_archs[:3],
                    multi=True,
                    style={"color": "#000"},
                ),
            ], md=4),
            dbc.Col([
                html.Label("Run Group (optional)", className="text-muted"),
                dcc.Dropdown(
                    id="lc-run-group",
                    options=[{"label": "All", "value": "all"}] +
                            [{"label": RUN_FOLDER_MAP.get(rg, rg), "value": rg} for rg in all_run_groups],
                    value="all",
                    style={"color": "#000"},
                ),
            ], md=2),
            dbc.Col([
                html.Label("Config Index (optional)", className="text-muted"),
                dcc.Dropdown(id="lc-config", options=[], value=None, style={"color": "#000"}),
            ], md=2),
            dbc.Col([
                html.Label("Smoothing Window", className="text-muted"),
                dcc.Slider(id="lc-smoothing", min=1, max=200, step=1, value=50,
                           marks={1: "1", 50: "50", 100: "100", 200: "200"},
                           tooltip={"placement": "bottom"}),
            ], md=2),
            dbc.Col([
                html.Label("Individual Runs", className="text-muted"),
                dbc.Checklist(
                    id="lc-individual",
                    options=[{"label": " Show", "value": "show"}],
                    value=[],
                    className="mt-2",
                ),
            ], md=2),
        ], className="mb-3"),
        dbc.Row([
            dbc.Col(dcc.Graph(id="lc-chart"), md=12),
        ]),
    ], fluid=True, className="py-3")

    # -----------------------------------------------------------------------
    # Loss Analysis Tab
    # -----------------------------------------------------------------------
    loss_tab = dbc.Container([
        dbc.Row([
            dbc.Col([
                html.Label("Architecture", className="text-muted"),
                dcc.Dropdown(
                    id="loss-arch",
                    options=[{"label": a, "value": a} for a in all_archs],
                    value=all_archs[0] if all_archs else None,
                    style={"color": "#000"},
                ),
            ], md=3),
            dbc.Col([
                html.Label("Run Group (optional)", className="text-muted"),
                dcc.Dropdown(
                    id="loss-run-group",
                    options=[{"label": "All", "value": "all"}] +
                            [{"label": RUN_FOLDER_MAP.get(rg, rg), "value": rg} for rg in all_run_groups],
                    value="all",
                    style={"color": "#000"},
                ),
            ], md=3),
            dbc.Col([
                html.Label("Config Index (optional)", className="text-muted"),
                dcc.Dropdown(id="loss-config", options=[], value=None, style={"color": "#000"}),
            ], md=3),
            dbc.Col([
                html.Label("Smoothing Window", className="text-muted"),
                dcc.Slider(id="loss-smoothing", min=1, max=200, step=1, value=50,
                           marks={1: "1", 50: "50", 100: "100", 200: "200"},
                           tooltip={"placement": "bottom"}),
            ], md=3),
        ], className="mb-3"),
        dbc.Row([
            dbc.Col(dcc.Graph(id="loss-components"), md=7),
            dbc.Col(dcc.Graph(id="loss-vs-reward"), md=5),
        ]),
    ], fluid=True, className="py-3")

    # -----------------------------------------------------------------------
    # Leaderboard Tab
    # -----------------------------------------------------------------------
    lb_cols = ["rank", "arch_label", "run_group", "config_idx", "run_idx",
               "avg_last_100", "best_reward", "mean_reward", "first_epoch_to_zero",
               "lr", "K_epochs", "epsilon", "gamma", "param_count"]
    lb_display = leaderboard[lb_cols].copy()
    lb_display.columns = ["Rank", "Architecture", "Run Group", "Config", "Run",
                          "Avg Last 100", "Best Reward", "Mean Reward", "First to 0",
                          "LR", "K Epochs", "Epsilon", "Gamma", "Params"]
    for col in ["Avg Last 100", "Best Reward", "Mean Reward"]:
        lb_display[col] = lb_display[col].round(2)
    lb_display["First to 0"] = lb_display["First to 0"].apply(
        lambda x: f"{x:.0f}" if not np.isnan(x) else "N/A")

    leaderboard_tab = dbc.Container([
        dbc.Row([
            dbc.Col([
                html.Div([
                    html.H5("Experiment Leaderboard", className="text-muted d-inline"),
                    html.Button("Download CSV", id="lb-download-btn",
                                className="btn btn-outline-info btn-sm ms-3"),
                    dcc.Download(id="lb-download"),
                ], className="mb-3"),
                dash_table.DataTable(
                    id="lb-table",
                    data=lb_display.to_dict("records"),
                    columns=[{"name": c, "id": c} for c in lb_display.columns],
                    style_table={"overflowX": "auto", "maxHeight": "600px", "overflowY": "auto"},
                    style_header={
                        "backgroundColor": CARD_BG, "color": "white",
                        "fontWeight": "bold", "border": "1px solid #2a2a4a",
                        "position": "sticky", "top": 0,
                    },
                    style_cell={
                        "backgroundColor": DARK_BG, "color": "#e0e0e0",
                        "border": "1px solid #2a2a4a", "textAlign": "center",
                        "padding": "6px 10px", "minWidth": "80px",
                    },
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#1e1e3a"},
                    ],
                    sort_action="native",
                    filter_action="native",
                    page_size=50,
                ),
            ], md=12),
        ]),
    ], fluid=True, className="py-3")

    # -----------------------------------------------------------------------
    # Main Layout
    # -----------------------------------------------------------------------
    app.layout = dbc.Container([
        dbc.Row([
            dbc.Col([
                html.H2("PPO Flappy Bird Training Analysis",
                         style={"color": ACCENT, "fontWeight": "bold"}),
                html.P(f"{kpis['total_experiments']} experiments across 5 hyperparameter sweeps | 10 architectures",
                        className="text-muted mb-0", style={"fontSize": "0.9rem"}),
            ], className="py-3"),
        ]),
        dbc.Tabs([
            dbc.Tab(overview_tab, label="Overview", tab_id="tab-overview"),
            dbc.Tab(perrun_tab, label="Per-Run Analysis", tab_id="tab-perrun"),
            dbc.Tab(model_tab, label="Model Comparison", tab_id="tab-model"),
            dbc.Tab(lc_tab, label="Learning Curves", tab_id="tab-lc"),
            dbc.Tab(loss_tab, label="Loss Analysis", tab_id="tab-loss"),
            dbc.Tab(leaderboard_tab, label="Leaderboard", tab_id="tab-lb"),
        ], id="tabs", active_tab="tab-overview", className="mb-0"),
    ], fluid=True, style={"backgroundColor": DARK_BG, "minHeight": "100vh"})

    # -----------------------------------------------------------------------
    # Callbacks
    # -----------------------------------------------------------------------

    # Per-Run: update charts when run group changes
    @app.callback(
        [Output("perrun-sweep", "figure"),
         Output("perrun-heatmap", "figure"),
         Output("perrun-box", "figure")],
        [Input("perrun-run-group", "value")],
    )
    def update_perrun(run_group):
        if not run_group:
            empty = dark_layout(go.Figure(), "Select a run group")
            return empty, empty, empty
        return (
            fig_hyperparam_sweep(summary, run_group),
            fig_heatmap(summary, run_group),
            fig_run_box(summary, run_group),
        )

    # Learning Curves: update config dropdown when run group changes
    @app.callback(
        [Output("lc-config", "options"), Output("lc-config", "value")],
        [Input("lc-run-group", "value")],
    )
    def update_lc_config_options(run_group):
        if not run_group or run_group == "all":
            configs = sorted(df["config_idx"].unique())
        else:
            configs = sorted(df[df["run_group"] == run_group]["config_idx"].unique())
        opts = [{"label": "All", "value": "all"}] + [{"label": f"Config {c}", "value": c} for c in configs]
        return opts, "all"

    # Learning Curves: main chart
    @app.callback(
        Output("lc-chart", "figure"),
        [Input("lc-archs", "value"),
         Input("lc-run-group", "value"),
         Input("lc-config", "value"),
         Input("lc-smoothing", "value"),
         Input("lc-individual", "value")],
    )
    def update_lc(archs, run_group, config_idx, smoothing, individual):
        if not archs:
            return dark_layout(go.Figure(), "Select at least one architecture")
        rg = None if run_group == "all" else run_group
        ci = None if config_idx == "all" or config_idx is None else int(config_idx)
        show_ind = "show" in (individual or [])
        return fig_learning_curves(df, archs, rg, ci, smoothing or 50, show_ind)

    # Loss Analysis: update config dropdown when run group changes
    @app.callback(
        [Output("loss-config", "options"), Output("loss-config", "value")],
        [Input("loss-run-group", "value")],
    )
    def update_loss_config_options(run_group):
        if not run_group or run_group == "all":
            configs = sorted(df["config_idx"].unique())
        else:
            configs = sorted(df[df["run_group"] == run_group]["config_idx"].unique())
        opts = [{"label": "All", "value": "all"}] + [{"label": f"Config {c}", "value": c} for c in configs]
        return opts, "all"

    # Loss Analysis: main charts
    @app.callback(
        [Output("loss-components", "figure"),
         Output("loss-vs-reward", "figure")],
        [Input("loss-arch", "value"),
         Input("loss-run-group", "value"),
         Input("loss-config", "value"),
         Input("loss-smoothing", "value")],
    )
    def update_loss(arch, run_group, config_idx, smoothing):
        if not arch:
            empty = dark_layout(go.Figure(), "Select an architecture")
            return empty, empty
        rg = None if run_group == "all" else run_group
        ci = None if config_idx == "all" or config_idx is None else int(config_idx)
        return (
            fig_loss_components(df, arch, rg, ci, smoothing or 50),
            fig_loss_vs_reward(df, arch, rg, ci),
        )

    # Leaderboard CSV download
    @app.callback(
        Output("lb-download", "data"),
        [Input("lb-download-btn", "n_clicks")],
        prevent_initial_call=True,
    )
    def download_csv(n_clicks):
        return dcc.send_data_frame(lb_display.to_csv, "ppo_flappy_leaderboard.csv", index=False)

    return app


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Loading data...")
    df = load_all_data()

    if df.empty:
        print("ERROR: No data loaded. Check that Excel files exist in the run folders.")
        exit(1)

    print("Computing metrics...")
    summary, kpis, arch_comp, leaderboard = compute_all(df)

    print(f"\n{'='*60}")
    print(f"  Best reward:    {kpis['best_reward']:.1f}  ({kpis['best_reward_id']})")
    print(f"  Best avg:       {kpis['best_avg']:.1f}  ({kpis['best_avg_id']})")
    print(f"  Best arch:      {kpis['best_arch']}  (avg={kpis['best_arch_avg']:.1f})")
    print(f"  Success rate:   {kpis['success_rate']:.0f}%")
    print(f"  Experiments:    {kpis['total_experiments']}")
    print(f"{'='*60}\n")

    print("Building dashboard...")
    app = build_app(df, summary, kpis, arch_comp, leaderboard)

    print("Starting server at http://localhost:8050")
    app.run(debug=True, port=8050)
